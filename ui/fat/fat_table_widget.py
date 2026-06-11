"""ui/fat/fat_table_widget.py"""
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QLineEdit,
    QComboBox, QAbstractItemView, QFrame, QMessageBox, QMenu,
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui  import QColor, QFont, QAction

from config.settings       import THEME as T, FAT_STATUSES, STATUS_COLORS
from services.fat_service  import FATService
from repositories.project_repository  import ProjectRepository
from repositories.supplier_repository import SupplierRepository

COLUMNS = [
    ("serial_no",           "#",            50),
    ("project",             "Project",      160),
    ("supplier",            "Supplier",     160),
    ("inspection_date",     "Date",         100),
    ("description",         "Description",  220),
    ("quantity",            "Qty",          60),
    ("status",              "Status",       160),
    ("supplier_contact",    "Supplier Eng.",130),
    ("consultant",          "Consultant",   130),
    ("po_no",               "PO No",        120),
    ("sap_no",              "SAP No",       100),
    ("major_comments",      "Comments",     200),
]


class FATTableWidget(QWidget):
    def __init__(self, session, current_user, parent=None):
        super().__init__(parent)
        self.session      = session
        self.current_user = current_user
        self._all_records = []
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self._apply_search)
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 18, 20, 18)
        root.setSpacing(14)

        # ── Header row ────────────────────────────────────────────────────
        header_row = QHBoxLayout()
        title = QLabel("FAT Records")
        title.setStyleSheet(f"font-size: 22px; font-weight: bold; color: {T['primary']};")
        header_row.addWidget(title)
        header_row.addStretch()

        # Search
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍  Search records…")
        self.search_input.setObjectName("search_bar")
        self.search_input.setFixedWidth(300)
        self.search_input.textChanged.connect(lambda: self._search_timer.start(300))
        header_row.addWidget(self.search_input)

        add_btn = QPushButton("＋  Add FAT Record")
        add_btn.setStyleSheet(f"background-color:{T['primary']}; color:white; font-weight:bold; padding:0 16px; border-radius:4px;")
        add_btn.setFixedHeight(34)
        add_btn.clicked.connect(self._add_record)
        header_row.addWidget(add_btn)

        export_btn = QPushButton("⬇  Export")
        export_btn.setFixedHeight(36)
        export_btn.setStyleSheet("background:#27ae60; color:white; border-radius:5px; padding: 0 14px;")
        export_btn.clicked.connect(self._export_excel)
        header_row.addWidget(export_btn)
        root.addLayout(header_row)

        # ── Filter bar ────────────────────────────────────────────────────
        filter_frame = QFrame()
        filter_frame.setStyleSheet(f"background:white; border-radius:8px; border:1px solid {T['border']};")
        filter_lay = QHBoxLayout(filter_frame)
        filter_lay.setContentsMargins(12, 8, 12, 8)
        filter_lay.setSpacing(10)

        filter_lbl = QLabel("Filter:")
        filter_lbl.setStyleSheet(f"color:{T['text_light']}; font-size:12px; font-weight:bold;")
        filter_lay.addWidget(filter_lbl)

        self.project_filter = QComboBox()
        self.project_filter.setFixedWidth(180)
        self.project_filter.addItem("All Projects", None)
        self.project_filter.currentIndexChanged.connect(self._apply_filter)

        self.supplier_filter = QComboBox()
        self.supplier_filter.setFixedWidth(180)
        self.supplier_filter.addItem("All Suppliers", None)
        self.supplier_filter.currentIndexChanged.connect(self._apply_filter)

        self.status_filter = QComboBox()
        self.status_filter.setFixedWidth(180)
        self.status_filter.addItem("All Statuses", None)
        for s in FAT_STATUSES:
            self.status_filter.addItem(s, s)
        self.status_filter.currentIndexChanged.connect(self._apply_filter)

        filter_lay.addWidget(self.project_filter)
        filter_lay.addWidget(self.supplier_filter)
        filter_lay.addWidget(self.status_filter)
        filter_lay.addStretch()

        self.count_lbl = QLabel("0 records")
        self.count_lbl.setStyleSheet(f"color:{T['text_light']}; font-size:12px;")
        filter_lay.addWidget(self.count_lbl)

        reset_btn = QPushButton("Reset Filters")
        reset_btn.setFixedHeight(28)
        reset_btn.setStyleSheet("background:#95a5a6; color:white; border-radius:4px; padding:0 10px; font-size:11px;")
        reset_btn.clicked.connect(self._reset_filters)
        filter_lay.addWidget(reset_btn)
        root.addWidget(filter_frame)

        # ── Table ─────────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(len(COLUMNS))
        self.table.setHorizontalHeaderLabels([c[1] for c in COLUMNS])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._context_menu)
        self.table.doubleClicked.connect(self._open_detail)

        for i, (_, _, w) in enumerate(COLUMNS):
            self.table.setColumnWidth(i, w)

        root.addWidget(self.table)

    # ── Data ──────────────────────────────────────────────────────────────
    def refresh(self):
        self._load_filters()
        svc = FATService(self.session)
        self._all_records = svc.get_all()
        self._populate_table(self._all_records)

    def _load_filters(self):
        proj_repo = ProjectRepository(self.session)
        sup_repo  = SupplierRepository(self.session)

        self.project_filter.blockSignals(True)
        self.supplier_filter.blockSignals(True)
        cur_proj = self.project_filter.currentData()
        cur_sup  = self.supplier_filter.currentData()

        self.project_filter.clear()
        self.project_filter.addItem("All Projects", None)
        for p in proj_repo.get_all_ordered():
            self.project_filter.addItem(p.project_name, p.id)

        self.supplier_filter.clear()
        self.supplier_filter.addItem("All Suppliers", None)
        for s in sup_repo.get_all_ordered():
            self.supplier_filter.addItem(s.supplier_name, s.id)

        # Restore selection
        for i in range(self.project_filter.count()):
            if self.project_filter.itemData(i) == cur_proj:
                self.project_filter.setCurrentIndex(i)
        for i in range(self.supplier_filter.count()):
            if self.supplier_filter.itemData(i) == cur_sup:
                self.supplier_filter.setCurrentIndex(i)

        self.project_filter.blockSignals(False)
        self.supplier_filter.blockSignals(False)

    def _populate_table(self, records):
        self.table.setRowCount(0)
        self.table.setRowCount(len(records))

        for row_idx, rec in enumerate(records):
            self.table.setRowHeight(row_idx, 36)
            values = {
                "serial_no":       str(rec.serial_no or ""),
                "project":         rec.project.project_name if rec.project else "",
                "supplier":        rec.supplier.supplier_name if rec.supplier else "",
                "inspection_date": str(rec.inspection_date or ""),
                "description":     rec.description or "",
                "quantity":        rec.quantity or "",
                "status":          rec.status or "",
                "supplier_contact": rec.supplier_contact.name if rec.supplier_contact else "",
                "consultant":      rec.consultant.company_name if rec.consultant else "",
                "po_no":           rec.po_no or "",
                "sap_no":          rec.sap_no or "",
                "major_comments":  rec.major_comments or "",
            }

            for col_idx, (key, _, _) in enumerate(COLUMNS):
                item = QTableWidgetItem(values.get(key, ""))
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                item.setData(Qt.ItemDataRole.UserRole, rec.id)

                # Color-code status column
                if key == "status":
                    from ui.shared.status_badge import StatusBadge
                    badge = StatusBadge(rec.status or "")
                    
                    # We wrap it in a QWidget layout to center it and prevent it from filling the whole cell
                    container = QWidget()
                    lay = QHBoxLayout(container)
                    lay.setContentsMargins(4, 2, 4, 2)
                    lay.addWidget(badge)
                    lay.addStretch()
                    
                    self.table.setCellWidget(row_idx, col_idx, container)
                    item.setText("") # Clear text since widget takes over
                
                self.table.setItem(row_idx, col_idx, item)

        self.count_lbl.setText(f"{len(records)} record{'s' if len(records) != 1 else ''}")

    # ── Search & Filter ───────────────────────────────────────────────────
    def _apply_search(self):
        keyword = self.search_input.text().strip()
        if not keyword:
            self._apply_filter()
            return
        svc     = FATService(self.session)
        results = svc.search(keyword)
        self._populate_table(results)

    def _apply_filter(self):
        if self.search_input.text().strip():
            return
        svc = FATService(self.session)
        results = svc.filter(
            project_id  = self.project_filter.currentData(),
            supplier_id = self.supplier_filter.currentData(),
            status      = self.status_filter.currentData(),
        )
        self._populate_table(results)

    def _reset_filters(self):
        self.search_input.clear()
        self.project_filter.setCurrentIndex(0)
        self.supplier_filter.setCurrentIndex(0)
        self.status_filter.setCurrentIndex(0)
        self._populate_table(self._all_records)

    # ── CRUD actions ──────────────────────────────────────────────────────
    def _add_record(self):
        from ui.fat.fat_form_dialog import FATFormDialog
        dlg = FATFormDialog(self.session, self.current_user, parent=self)
        if dlg.exec():
            self.refresh()

    def _open_detail(self):
        row = self.table.currentRow()
        if row < 0:
            return
        rec_id = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        from ui.fat.fat_detail_dialog import FATDetailDialog
        dlg = FATDetailDialog(self.session, self.current_user, record_id=rec_id, parent=self)
        if dlg.exec():
            self.refresh()

    def _edit_selected(self):
        row = self.table.currentRow()
        if row < 0:
            return
        rec_id = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)
        from ui.fat.fat_form_dialog import FATFormDialog
        dlg = FATFormDialog(self.session, self.current_user, record_id=rec_id, parent=self)
        if dlg.exec():
            self.refresh()

    def _context_menu(self, pos):
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        rec_id = self.table.item(row, 0).data(Qt.ItemDataRole.UserRole)

        menu = QMenu(self)
        edit_act       = menu.addAction("✏️  Edit")
        duplicate_act  = menu.addAction("⧉  Duplicate")
        menu.addSeparator()
        delete_act     = menu.addAction("🗑  Delete")

        action = menu.exec(self.table.viewport().mapToGlobal(pos))

        if action == edit_act:
            from ui.fat.fat_form_dialog import FATFormDialog
            dlg = FATFormDialog(self.session, self.current_user, record_id=rec_id, parent=self)
            if dlg.exec():
                self.refresh()

        elif action == duplicate_act:
            svc = FATService(self.session, self.current_user.id)
            svc.duplicate(rec_id)
            self.refresh()

        elif action == delete_act:
            if self.current_user.role != "Admin":
                QMessageBox.warning(self, "Permission Denied", "Only Admin can delete records.")
                return
            reply = QMessageBox.question(
                self, "Confirm Delete",
                "Are you sure you want to delete this record?\nThis action cannot be undone.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.Yes:
                svc = FATService(self.session, self.current_user.id)
                svc.delete(rec_id)
                self.refresh()

    def _export_excel(self):
        from PyQt6.QtWidgets import QFileDialog
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from config.settings import STATUS_COLORS
        import datetime

        path, _ = QFileDialog.getSaveFileName(
            self, "Export to Excel",
            f"FAT_Records_{datetime.date.today()}.xlsx",
            "Excel Files (*.xlsx)",
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "FAT Records"

        headers = [c[1] for c in COLUMNS]
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="1A3A5C")
            cell.alignment = Alignment(horizontal="center")

        for row_idx in range(self.table.rowCount()):
            row_data = [
                self.table.item(row_idx, col).text() if self.table.item(row_idx, col) else ""
                for col in range(self.table.columnCount())
            ]
            ws.append(row_data)

        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(path)
        QMessageBox.information(self, "Export Complete", f"File saved:\n{path}")
