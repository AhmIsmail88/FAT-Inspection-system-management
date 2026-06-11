"""ui/reports/reports_widget.py"""
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QLabel, QPushButton, QHBoxLayout, QComboBox, QFileDialog, QMessageBox
from PyQt6.QtCore    import Qt
from config.settings import THEME as T


class ReportsWidget(QWidget):
    def __init__(self, session, current_user, parent=None):
        super().__init__(parent)
        self.session = session
        self.current_user = current_user
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(40, 30, 40, 30)
        root.setSpacing(20)

        title = QLabel("Reports")
        title.setStyleSheet(f"font-size:22px;font-weight:bold;color:{T['primary']};")
        root.addWidget(title)

        cards_lay = QHBoxLayout()
        cards_lay.setSpacing(16)

        report_types = [
            ("📊 Full FAT Report",    "All records with filters",     "#2980b9", self._export_full),
            ("🏗️  Project Report",   "All FATs for a project",        "#27ae60", self._export_project),
            ("🏭  Supplier Report",   "All FATs for a supplier",       "#8e44ad", self._export_supplier),
            ("📅  Monthly Report",    "Records grouped by month",      "#e67e22", self._export_monthly),
        ]

        for title_text, desc, color, handler in report_types:
            card = self._make_report_card(title_text, desc, color, handler)
            cards_lay.addWidget(card)

        root.addLayout(cards_lay)
        root.addStretch()

        note = QLabel("Reports are exported as Excel (.xlsx) files with company branding.")
        note.setStyleSheet(f"color:{T['text_light']}; font-size:12px;")
        note.setAlignment(Qt.AlignmentFlag.AlignCenter)
        root.addWidget(note)

    def _make_report_card(self, title, desc, color, handler):
        from PyQt6.QtWidgets import QFrame
        card = QFrame()
        card.setFixedSize(220, 160)
        card.setStyleSheet(f"""
            QFrame {{ background:white; border-radius:10px;
                      border-left: 5px solid {color};
                      border-top:1px solid #e0e0e0;
                      border-right:1px solid #e0e0e0;
                      border-bottom:1px solid #e0e0e0; }}
        """)
        lay = QVBoxLayout(card)
        lay.setContentsMargins(16, 16, 16, 16)

        t = QLabel(title)
        t.setStyleSheet(f"font-size:13px;font-weight:bold;color:{T['primary']};border:none;")
        lay.addWidget(t)

        d = QLabel(desc)
        d.setStyleSheet(f"font-size:11px;color:{T['text_light']};border:none;")
        d.setWordWrap(True)
        lay.addWidget(d)
        lay.addStretch()

        btn = QPushButton("Export Excel")
        btn.setStyleSheet(f"background:{color};color:white;border-radius:5px;padding:6px;border:none;")
        btn.clicked.connect(handler)
        lay.addWidget(btn)
        return card

    def refresh(self): pass

    def _get_save_path(self, default_name: str) -> str | None:
        path, _ = QFileDialog.getSaveFileName(self, "Save Report", default_name, "Excel Files (*.xlsx)")
        return path if path else None

    def _export_full(self):
        path = self._get_save_path("FAT_Full_Report.xlsx")
        if path: self._write_excel(path, "Full Report")

    def _export_project(self):
        path = self._get_save_path("FAT_Project_Report.xlsx")
        if path: self._write_excel(path, "Project Report")

    def _export_supplier(self):
        path = self._get_save_path("FAT_Supplier_Report.xlsx")
        if path: self._write_excel(path, "Supplier Report")

    def _export_monthly(self):
        path = self._get_save_path("FAT_Monthly_Report.xlsx")
        if path: self._write_excel(path, "Monthly Report")

    def _write_excel(self, path: str, report_title: str):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from services.fat_service import FATService
        import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "FAT Records"

        # Header row
        headers = ["#","Project","Supplier","Date","Description","Qty","Status","Supplier Eng.","Consultant","PO No","SAP No","Comments"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor="1A3A5C")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        svc = FATService(self.session)
        for rec in svc.get_all():
            ws.append([
                rec.serial_no,
                rec.project.project_name if rec.project else "",
                rec.supplier.supplier_name if rec.supplier else "",
                str(rec.inspection_date or ""),
                rec.description or "",
                rec.quantity or "",
                rec.status or "",
                rec.supplier_contact.name if rec.supplier_contact else "",
                rec.consultant.company_name if rec.consultant else "",
                rec.po_no or "",
                rec.sap_no or "",
                rec.major_comments or "",
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 4, 50
            )

        # Watermark row at bottom
        ws.append([])
        ws.append([f"HAC FAT System — {report_title} — Generated: {datetime.date.today()}"])

        wb.save(path)
        QMessageBox.information(self, "Export Complete", f"Report saved:\n{path}")
